# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html


# useful for handling different item types with a single interface

from openpyxl import Workbook
from .items import *

class ChinalocationPipeline(object):
    def open_spider(self, Province):
        # 当爬虫开启时，创建 Excel 文件和工作表
        self.workbook = Workbook()
        self.sheet_a = self.workbook.active
        self.sheet_1 = self.workbook.create_sheet(title='省')
        self.sheet_2 = self.workbook.create_sheet(title='市')
        self.sheet_3 = self.workbook.create_sheet(title='县区')
        self.sheet_4 = self.workbook.create_sheet(title='Class4_Item')
        self.sheet_5 = self.workbook.create_sheet(title='Class5_Item')

        self.sheet_1.append(['code', 'name'])  # Add headers for ItemA
        self.sheet_2.append(['code', 'name'])  # Add headers for ItemB
        self.sheet_3.append(['code', 'name'])  # 添加表头
        self.sheet_4.append(['code', 'name'])  # 添加表头
        self.sheet_5.append(['code','code2', 'name'])  # 添加表头

    def process_item(self, item, Province):
        # 处理每个 Item，并将数据写入工作表
        # self.worksheet.append([item['code'], item['name']])

        if isinstance(item, Class1_Item):
            self.sheet_1.append([item['code'], item['name']])
        elif isinstance(item, Class2_Item):
            self.sheet_2.append([item['code'], item['name']])
        elif isinstance(item, Class3_Item):
            self.sheet_3.append([item['code'], item['name']])
        elif isinstance(item, Class4_Item):
            self.sheet_4.append([item['code'], item['name']])
        elif isinstance(item, Class5_Item):
            self.sheet_5.append([item['code'],item['code2'], item['name']])
        return item

    def close_spider(self, Province):
        # 当爬虫关闭时，保存 Excel 文件
        self.workbook.save('Province.xlsx')
